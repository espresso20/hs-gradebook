#!/usr/bin/env python3
import sys

# Try different libraries to read the Excel file
try:
    import pandas as pd
    print("Using pandas")
    
    # Read all sheets
    excel_file = pd.ExcelFile('fivejs-gradebook-plus-v1-excel.xlt')
    print(f"\nSheet names: {excel_file.sheet_names}")
    
    for sheet_name in excel_file.sheet_names:
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print('='*60)
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        print(f"\nShape: {df.shape}")
        print(f"\nColumns: {list(df.columns)}")
        print(f"\nFirst few rows:")
        print(df.head(10))
        print(f"\nData types:")
        print(df.dtypes)
        
except ImportError as e1:
    print(f"pandas not available: {e1}")
    try:
        import openpyxl
        print("Using openpyxl")
        
        wb = openpyxl.load_workbook('fivejs-gradebook-plus-v1-excel.xlt')
        print(f"\nSheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*60}")
            print(f"Sheet: {sheet_name}")
            print('='*60)
            ws = wb[sheet_name]
            print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
            
            # Print first 10 rows
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i <= 10:
                    print(row)
                else:
                    break
                    
    except ImportError as e2:
        print(f"openpyxl not available: {e2}")
        print("\nPlease install required libraries:")
        print("pip install pandas openpyxl")
        sys.exit(1)
